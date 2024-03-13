"""
openpyxl模块是用于处理Microsoft Excel文件的第三方库
可以对Excel文件中的数据进行写入和读取
函数/属性名称                             功能描述
load_workbook(filename)                 打开已存在的表格，结果为工作簿对象
workbook.sheetnames                     工作簿对象的sheetnames属性，用于获取所有工作表的名称，结果为列表类型
sheet.append(lst)                       向工作表中添加一行数据，新数据接在工作表已有数据的后面
workbook.save(excelname)                保存工作簿
Workbook()                              创建新的工作簿对象
"""
import requests
import re
import openpyxl


# 定义函数
def get_html():
    url = 'http://www.weather.com.cn/weather1d/101230201.shtml'
    resp = requests.get(url)
    # 设置一下编码模式
    resp.encoding = 'utf-8'
    return resp.text


def parse_html(html_str):
    city = re.findall('<span class="name">([\u4e00-\u9fa5]*)</span>', html_str)
    weather = re.findall('<span class="weather">([\u4e00-\u9fa5]*)</span>', html_str)
    wd = re.findall('<span class="wd">(.*)</span>', html_str)
    zs = re.findall('<span class="zs">([\u4e00-\u9fa5]*)</span>', html_str)
    lst = []
    for a, b, c, d in zip(city, weather, wd, zs):
        lst.append([a, b, c, d])
    return lst


# 从爬虫处获取数据
html = get_html()
lst = parse_html(html)
# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()  # 创建工作簿对象
# 在Excel文件中创建工作表
sheet = workbook.create_sheet('景区天气')
# 向工作表中加数据
for item in lst:
    sheet.append(item)
workbook.save('景区天气.xlsx')

# 从Excel中读取数据
workbook = openpyxl.load_workbook('景区天气.xlsx')
# 选择要操作的工作表
sheet = workbook['景区天气']
# 表格数据是二维列表，先遍历的是行，后遍历的是列
lst = []  # 存储的是行数据
for row in sheet.rows:
    sublst = []  # 存储单元格数据
    for cell in row:
        sublst.append(cell.value)
    lst.append(sublst)
for item in lst:
    print(item)
